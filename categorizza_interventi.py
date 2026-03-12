"""
Categorizzazione interventi di manutenzione.
Legge manutenzione.xlsx, classifica ogni attività in 15 macro-categorie,
produce catalogo_interventi.xlsx con catalogo + dettaglio.
"""

import re
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 15 macro-categorie con pattern di riconoscimento ──────────────────────────
# Ordine importante: le categorie più specifiche vanno prima di quelle generiche

CATEGORIE = [
    # (nome_categoria, pattern_regex)
    # Ordine: le categorie più specifiche prima di quelle generiche

    ("01. PNEUMATICI", (
        r"pneum|pneu|gomm[ae]|gome\b|goma\b|jinyu|pirelli|bridgestone|goodyear|sava\b|"
        r"goodrich|michelin|rappezzo|foratur|inversione.*ruot|rotazione.*ruot|"
        r"valvola pneum|ruota di scorta|ruota bucata|ruot[ae].*scorta|"
        r"mont.*ruota.*scorta|stacco.*ruota|cerchio|"
        r"385/65|315/80|295/80|295/60|385/55|315/70|315/60|"
        r"12[\.,]00|13[\.,]00|r22[\.,]?5|r19[\.,]?5|r17[\.,]?5|"
        r"montaggio.*pneum|smontaggio.*pneum|smont\..*mont.*pneum|"
        r"gomme|cambio ruota|it\s*90|itineris|trailer\s*90|"
        r"penumatic|pneuamatic|"  # typos
        r"sostituzione.*ruot|cam\s*gom|allineamento.*ruot|"
        r"riparaz\w*\s*penu|riparaz\w*\s*pneum|"
        r"cam\.\s*goma|"
        r"copertone|inversione.*cerch|cerch"
    )),
    ("02. IMPIANTO FRENANTE", (
        r"fren[io]|feni\b|pastigl|disco\s*fren|disci\b|dischi\s*\d.*ass|brake|diapress|torples|"
        r"cilindro\s*fren|cilindro.*membran|cilindro.*molla|registrazione\s*fren|cavo\s*freno|soffiett.*fren|"
        r"prova\s*fren|freno\s*a\s*mano|freno\s*a\s*tamburo|ganasc|prova.*rulli|"
        r"camera\s*fren|brake\s*chamber|zanche|manovella|manov\.|reggiman|"
        r"soffione|pinza\s*fren|regolat.*fren|caliper|"
        r"giunto.*iso|attac[hc][io]?\s*aria|spirale.*aria|gomini.*aria|tubo\s*spirale|"
        r"corr\.?\s*fren|modulat.*ebs|"
        r"sost.*dischi\s*\d|sost\s*feni|somtaggio\s*disci|"
        r"load.*sens.*relay|"
        r"rep\s*tubi\s*aria|tubi\s*aria|"
        r"rubinetto.*sfrenat|sfrenatur|"
        r"rettifica.*disc|"
        r"trailer.*control.*valve|"
        r"soffioen|"
        r"presa.*ebs|"
        r"cilindro.*tristop|tristop|"
        r"cartuccia.*coalescenz|"
        r"valvol[ae].*scar.*cond|"
        r"valvol[oa].*control.*pression|"
        r"champio|ecoair"
    )),
    ("03. SOSPENSIONI E AMMORTIZZATORI", (
        r"moll[ae]\s*aria|air\s*sp[ir]|ammort|amort|ammo\b|silent.?block|balestra|bal\.\s*comp|"
        r"sospensi|suspens|susp\b|sosp\b|boccol[ea]|piastre usura|rondell.*sospen|"
        r"cv\s*shock|assale|asale|barra\s*stabiliz|stabiliz[z]?at|tampone|tamponi|"
        r"moll[ae].*pneumat|moll[ae].*asse|spring\s*seat|"
        r"cam\s+susp|sost\s*susp|mont\s*susp|sost\s*sosp|sost\s*sospe|"
        r"cam\s+\d*\s*amort|cam\s*amort|sost\s*ammo|"
        r"serraggio\s*balest|"
        r"raise.*lower.*valve|levelling.*valve|"
        r"elemento\s*elastic|ammorizzat|"
        r"5\s*ammorizzat"
    )),
    ("04. CARROZZERIA CONTAINER / CASSE MOBILI", (
        # Pavimenti e pianali (anche typo PIOANALI)
        r"paviment|pianal[ei]?|pioanal|multistrato|betulla|tavol[eao]|"
        # Stecche (anche abbreviate e typo: STEQUE, STECHE, STECC)
        r"stecc[ha]|stecch|stequ|steche|"
        # Ruggine (anche typo RUIGGINE, RUGINE)
        r"ruggin|ruigg|rugine|"
        # Porte (blocco, interne, serratura)
        r"port[ea].*(?:blocc|intern|dx|sx|post)|sblocc?.*port|chiusur.*port|"
        r"porte\s*blocc|sblocc?.*asta|serrattur|serratur|"
        r"chiusur[ae]?\s*cass|"
        # Traverse (anche abbreviate: TRAV., typo RAVERSA)
        r"travers|trav\.|ravers[ae]|"
        # Cerniere (anche typo)
        r"cernier[ae]|"
        # Ganci (anche GANGIO typo, GANCETTO)
        r"ganc[io]|gancett|gangio|"
        # Perni
        r"pern[io](?!.*ruota)|"
        # Botole
        r"botol[ea]|"
        # Montanti e longheroni
        r"montant[ie]|longher|"
        # J-Bar
        r"j.?bar|"
        # Maniglie
        r"manigli[ae]|"
        # Corrimano
        r"corrimano|"
        # Ventole (container)
        r"ventol[ae]|"
        # Scaletta/scala
        r"scalett[ae]|scala\s*sal|"
        # Cancello
        r"cancell[oi]|"
        # Colpi laterali (anche abbreviato: COLPO LATO, COLP LAT, COLPO)
        r"colp[io].*lat|colp[io].*[ds]x|colp[io].*lato|\bcolp[io]\b|"
        # Raddrizzatura (anche abbreviata: RADDR., RADD., RAD.)
        r"raddrizzat|raddriz|raddr\.|radd[r]?\.|rad\.\s*asta|"
        # Buchi
        r"buc[ho].*tetto|buc[ho].*paret|buc[ho].*[ds]x|tetto.*buc[ho]|"
        # Pannelli (no catarifrangenti)
        r"pannell[io](?!.*catarifrang|.*reflex)|kit\s*pannello|"
        # Verniciatura (anche typo VERNICATURA, VERNICITUR)
        r"verniciatur|vernicatur|trattamento ruggine|antiruggine|"
        # Guarnizioni (anche typo GURNIZIONE)
        r"guarnizion|gurnizion|"
        # Casse/container
        r"cass[ea].*acciaio|cass[ea].*ast|cass[ea]\s*mobil|container|\bcont\.\b|"
        # Saldature specifiche container
        r"sald\w*.*montant|sald\w*.*cernier|sald\w*.*pern|"
        r"sald\w*.*stecca|sald\w*.*tetto|sald\w*.*ganc|"
        # Angolari e lamiere per casse
        r"angolar[ie]|lamier[ae].*cass|"
        # Porte posteriori, asta, aletta, arpione
        r"porte.*post|sede.*manigl|"
        r"arpion[ei]|guid[ae].*port|asta.*port|aletta|basculant|"
        # CSC container
        r"\bcsc\b|"
        # Ferma porte
        r"ferma.*port|"
        # Taglio parete
        r"taglio.*paret|paret[ie].*[ds]x|"
        # Preparazione/riparazione cassa
        r"preparaz\w*\s*cass|riparaz\w*\s*cass|prep\s*cass|"
        r"preparaz\w*\s*tetto|preparaione\s*tetto|"
        r"riparaz\w*\s*port[ae]|"
        # Centine
        r"centina|centin[ae]|"
        # Tinteggiatura
        r"tinteggiatur|"
        # Botte/botta (colpi)
        r"\bbott[ae]\b|\bbotte\b|"
        # Plastica interna
        r"plastica\s*intern|"
        # Perdita dentro container
        r"perdita.*dentro|"
        # Rinforzamento tetto
        r"rinforzam|"
        # Collari e portatubi/manichette
        r"collar[ie]|portamanichett|portatub|porta\s*tub|caseta\s*tub|"
        # Bucato laterale (buco)
        r"bucat[oi].*later|later.*bucat|"
        # Cassone
        r"cassone|"
        # Preparazione + chiusura generico
        r"preparazione.*chiusur|chiusura\b"
    )),
    ("05. TELONI E COPERTURE", (
        r"telon[ei]|pezz[ea].*telon|cucitur|"
        r"cop\.\s*cass|copertur[ae].*cass|coperchio.*cass|coperta.*cass|"
        r"cop\.\s*casa|cop\s+cass|cop\s+casa|copeta\s*cas|coperta\s*cas|"
        r"cavo tir|cavett.*(?:nylon|tir)|"
        r"barra avvolgimento|"
        r"archett[io]|"
        r"tiparaz|rip\w*.*telon|"
        r"\bcricchett|cinghia.*cricchett|"
        r"sbloccaggio.*copertur|copertura\b|sacco\b|"
        r"teli\s*copri|telo\b|topp[ea].*tel|rip\w*.*telo|"
        r"riparaz\w*\s*telo"
    )),
    ("06. IMPIANTO ELETTRICO E LUCI", (
        r"\bluc[ie]\b|lucii|fanal[ei]|fanalin|lampad|cablaggio|"
        r"pres[ea].*elettr|pres[ea].*poli|pres[ea].*2p|presa\s*rimorchio|"
        r"24v|"
        r"spirale\s*elettr|spirale\s*plast|spirale\s*eletr|"
        r"interruttore|centralina.*frecc|"
        r"massa\s*a\s*terra|messa\s*a\s*terra|catarifrang|rifletten|"
        r"imp\.?\s*luci|eutopoint|ecopoint|"
        r"led\s|led\b|bianco.*led|arancio.*led|"
        r"pannell[io].*catarifrang|paneli?\s*reflex|panel\s*reflex|"
        r"imp\w*\s*elettr|elettric|eletric|electri|elec\b|"
        r"strisc[ie].*ross|segnalet|"
        r"rev\w*\s*elettr|rev\s*sist\s*electr|rep\s*sist.*elet|"
        r"gemma|portatarga|targa\s*ripetit|"
        r"videocamer|alimentator[ei]|"
        r"rele\b|rel[eè]|fusibil|"
        r"selettore|prossimit|"
        r"cavo.*scarico.*terra|cavo\s*messa|"
        r"spazzol[ea].*tergicristall|tergicristall|"
        r"collegam\w*.*cavi|montaggio.*alimentat|"
        r"filo\s*rimorchi|attacco\s*corrent|"
        r"spina\b|tabbelle|"
        r"montaggio\s*panelli.*luc|"
        r"spirale\s*ebs|"
        r"europoint|gemm[ae]|"
        r"strisci[ae]|"
        r"power\s*relay"
    )),
    ("07. MOTORE E MECCANICA MOTRICE", (
        r"tagliand|filtr[io].*olio|filtr[io].*carburant|filtr[io].*aria(?!.*silos)|"
        r"filtr[io].*antipolline|filtri\b|sost\s*filtr|\bfiltro\b|"
        r"cambio olio|perdita olio|rabbocco\s*olio|olio\s*atf|"
        r"guarnizion.*copp|guarnizion.*test|"
        r"sensor[ei]|kitas|"
        r"rigenerazion|catalizzat|"
        r"bulb[io].*olio|"
        r"additivo|"
        r"blocchetto.*accension|"
        r"servomaster|frizion[ei]|"
        r"travaso|travasi|"
        r"depotenziam|"
        r"controllo.*a/?c|condizionat|"
        r"differenzial|"
        r"motore|turbina|cinghia(?!.*cricchett)|cinghie\s*dentat|radiator|ventola\s*mot|"
        r"diagnos[it]|centralin[ae](?!.*frecc)|"
        r"avviamento|alternator|compressor[ei](?!.*silos)|"
        r"marmitta|scappamento|tubo.*gas|"
        r"coppa olio|asta olio|liquido refrig|antigelo|"
        r"pompa gasolio|pompa acqua|pompa\s*idraul|iniett|"
        r"shunt.*valve|valvola.*marc|"
        r"puleggi[ae]|cint[ea]\b|distributore(?!.*silos)|"
        r"olio.*riduttor|riduttore|"
        r"pistola\s*aria|pistola\s*soffiag|soffiaggio\s*cabin|"
        r"silenziator[ei]|raccord[io].*sprint|"
        r"presa.*di.*forza|albero.*presa|canna.*olio|tubo.*olio|tubazion.*olio|"
        r"tubo.*aspiraz|tubo.*scaric(?!.*silos)|tubi\s*scaric|"
        r"aria\s*scania|"
        r"scarico\s*memoria|memoria\s*di\s*massa|"
        r"sost\s*olio\s*comp|sost\s*olio\b|sost\s*tubi\b|"
        r"testine|flang[ia]|"
        r"paraolio|cam\s*paraolio|"
        r"controllo\s*cambio|cambio\s*attacco\s*olio|olio\s*cambio|"
        r"attacco\s*aria\s*abs|"
        r"barra\s*sterzo|sterzo|"
        r"specchio.*retrovis|retrovis|"
        r"pompa\s*alza|alza\s*cabin|"
        r"cooling\s*sis|"
        r"blocco\s*trazion|"
        r"convertitor|"
        r"motore?\s*idraul|danfoss|"
        r"cinghie\s*dentel|"
        r"organi\s*di\s*comando|"
        r"6\s*pk|"
        r"repair\s*kit|"
        r"manutenzioni\s*vari"
    )),
    ("08. MOZZI E RUOTE", (
        r"mozzo|mozzi|muzzo|colonn[ea]tt[ea]|colonn?in[ea]|dad[io].*ruot|dadi\b|"
        r"perno ruota|disco limitat|"
        r"cuscinett|cucinett?i|registrazione\s*mozz|registro\s*cuci|"
        r"co\.\s*c.*ruota|coppia.*zanche|"
        r"hub\s*assembl|"
        r"cam\s*colonin|sost\s*\d*\s*colon[en]|"
        r"rotocamera|cam\s*rotocam"
    )),
    ("09. REVISIONE E CONTROLLI PERIODICI", (
        r"revision[ei]|revizione|movimentazione.*revision|"
        r"check.*list|preparaz.*revision|"
        r"calibrazione|tachigrafo|dtco|cronotachigraf|"
        r"prova fumi|"
        r"tabell[ea].*adesiv|adesiv[io].*normat|"
        r"prova.*usl|"
        r"collaudo|"
        r"dichiarazione.*conform|libretto.*manutenzione|"
        r"linea\s*vita|conformit|"
        r"prerevis|pre.?revis|pre.?revizion|"
        r"rev\s*pressio|rev\s*telaio|"
        r"controlli\s*pre.?revis|"
        r"verifica\s*periodic|recipiente\s*a\s*pression|"
        r"controllo\s*funzional"
    )),
    ("10. ATTREZZATURE SILOS / CISTERNA", (
        r"silos|cistern|"
        r"tubo.*scarico|tubo.*mandata|"
        r"filtro.*aria.*silos|filtro.*aria.*compressor|"
        r"valvola.*farfall|valvola.*europa|valvola.*stell|valvola.*sicurezz|"
        r"valvola.*livettat|valvola\s*universal|"
        r"storz|portagomma|"
        r"manometr|manomet\b|glicerin|"
        r"termometr|"
        r"coperchio.*doppio|delrin|coperchio\b|riraraz.*coperchi|"
        r"pompa.*ingranag|"
        r"nippli|nipplo|curv[ea].*inox|"
        r"maniglia.*valvola|valvola.*sirca|"
        r"volantino.*zincat|volantio|"
        r"convogliat|"
        r"soffiante|"
        r"DIN.?150|"
        r"rubinett|robineto|piatto.*scaric|"
        r"raccord[io](?!.*sprint)|"
        r"bonded|"
        r"ralla|"
        r"fascett[ae]|"
        r"ghier[ae]|bussola.*arrest|cannocchial|"
        r"calott[ea]|tiranti.*calott|"
        r"sollevator[ei]|soll\.\s*[ds]x|"
        r"portatub|porta\s*tub|"
        r"pistone.*ribalt|"
        r"innesto.*rapid|giunzione|giunto\b|"
        r"adattator[ei]|"
        r"o.ring|"
        r"sup\.\s*aria|"
        r"guarniz\s*semicircol|"
        r"piedini|"
        r"escort\b|"
        r"scarico\s*aria|"
        r"tramoggi[ae]|"
        r"valvola\s*a\s*farbal|farballa|"
        r"riduzione\s*zin[cg]at|"
        r"tappo\b|"
        r"gommino|"
        r"tubo\s*termoplast"
    )),
    ("11. ROTOCELLA E TWIST LOCK", (
        r"rotocell|rotocela|twist|twister|"
        r"pulsantier[ae]|"
        r"cavo.*pulsant|"
        r"blocchetto.*guida|"
        r"jost|menci"
    )),
    ("12. SOCCORSO E INTERVENTI FUORI SEDE", (
        r"officina.*mobile|"
        r"soccorso|"
        r"fuori.*sede|"
        r"trasfert|"
        r"presso.*client|c/o\s|novamont|fiorenzuola|"
        r"smontaggio.*zampe|"
        r"sul.*posto|in.*francia|hamecher|"
        r"tiro\s*gru|gru\s*ferar|"
        r"baselle|"
        r"messa\s*in\s*moto|"
        r"presso\s*vs\s*sede|"
        r"trasporto.*rottura"
    )),
    ("13. MATERIALI DI CONSUMO E FLUIDI", (
        r"adblue|ad\s*blue|repostare\s*ad|scarico\s*ad|olio\s*atf|"
        r"grass[oa]|graso|lubrificaz|ingrassag|"
        r"olio idraul|"
        r"olio.*compressor.*mobil|rarus|"
        r"olio.*motor.*shell|rimula|olio motore|"
        r"material[ei].*consumo|mat\.?\s*consumo|"
        r"batter[iy]|batteria|"
        r"manodopera|"
        r"viti.*rondell|"
        r"pulizia"
    )),
    ("14. STRUTTURA METALLICA E SALDATURE", (
        r"sagomati.*lamier|taglio.*laser|pressopieg|"
        r"tubolar[ie].*zincat|tubolar[ei]|"
        r"lamier[ae]|"
        r"anell[io].*acciaio|anell[io].*inox|"
        r"saldatur[ae]|saldat[oi]|soldatur|"
        r"sald\.(?!.*montant|.*cernier|.*pern|.*stecc|.*tetto|.*ganc)|"
        r"paragfanc|parafang|paraganghi|paraurti|paraschi|paracicl|"
        r"staff[ea]|staffa|"
        r"piattaform|pedan|"
        r"struttur.*metall|"
        r"fazzoletto|fazzoletti|"
        r"supporto|serbatoio|cassett[aio]|"
        r"capannin[ae]|anello|"
        r"barra.*paraincastr|barra\s*parainc|"
        r"frame.*mount.*bracket|"
        r"cassetta.*zinc|"
        r"cassetta.*attrez|"
        r"piastr[ae].*lamier|piastra\s*sagomat|piastra\s*rinforz|"
        r"cavallott|"
        r"piedi\b|sostituz.*piedi|"
        r"saldametall|"
        r"lastra\b|telarubbertex|"
        r"prolungh[ae]"
    )),
    ("15. ALLESTIMENTO E PERSONALIZZAZIONE", (
        r"kit.*satellit|installaz.*satellit|satellit|"
        r"scritt[ea].*cabin|"
        r"porta.*doc|"
        r"montaggio.*liner|liner\b|lainer\b|"
        r"adesiv[io].*letter|letter[ea].*adesiv|"
        r"messa.*su.*strada|"
        r"allestiment|"
        r"tasca.*porta|"
        r"cavo.*nero|"
        r"numero\b|sost.*numero|"
        r"semiaccoppi|"
        r"impianto.*as24|as24\b|"
        r"remix|oso\s*\d|"
        r"truck.*don\b|truck.*sahara|tubo.*aria.*truck|"
        r"montaggio\s*zampa|"
        r"cassetta\s*zinc|"
        r"gps\b|"
        r"scrite\b|"
        r"letter[ea]\s*mancan|numeri\b|numerazione|"
        r"targhett[ae]|"
        r"toll\s*collect"
    )),
]

# Pattern compilati
CATEGORIE_COMPILED = [
    (nome, re.compile(pattern, re.IGNORECASE))
    for nome, pattern in CATEGORIE
]


def pulisci_testo(testo):
    """Normalizza il testo rimuovendo rumore."""
    if not testo:
        return ""
    testo = str(testo).strip()
    if testo.lower() == "none":
        return ""
    # Rimuovi caratteri di controllo Excel
    testo = testo.replace("_x000D_", "").replace("\r", "")
    # Rimuovi prezzi (€ 190,00)
    testo = re.sub(r"€\s*[\d.,]+", "", testo)
    # Rimuovi date isolate (del 11/02/25, DEL 11/02 E 14/02/25)
    testo = re.sub(r"\bdel\s+\d{1,2}/\d{1,2}(?:/\d{2,4})?\b", "", testo, flags=re.IGNORECASE)
    testo = re.sub(r"\be\s+\d{1,2}/\d{1,2}(?:/\d{2,4})?\b", "", testo, flags=re.IGNORECASE)
    # Rimuovi numeri container (cont. 2421)
    testo = re.sub(r"\bcont\.?\s*\d+", "", testo, flags=re.IGNORECASE)
    # Rimuovi "NR" / "N." prefix con numero
    testo = re.sub(r"\bn\.?\s*r?\.?\s*(?=\d)", "", testo, flags=re.IGNORECASE)
    # Rimuovi matricole isolate (tipo 8015671680, 32431066703)
    testo = re.sub(r"\bmatricol[ae]?:?\s*[\w\d/\-]+", "", testo, flags=re.IGNORECASE)
    # Rimuovi codici fornitori (COD. xxx, COD xxx)
    testo = re.sub(r"\bcod\.?\s*[\w\d.]+", "", testo, flags=re.IGNORECASE)
    return testo.strip()


def is_solo_codice(testo):
    """Controlla se il testo è solo un codice numerico/seriale senza contenuto."""
    t = testo.strip()
    if not t:
        return True
    # Solo numeri e spazi
    if re.match(r"^[\d\s/\-\.]+$", t):
        return True
    # Codice tipo AZJ21136 R, HJT 03473 F 2DX
    if re.match(r"^[A-Z]{2,4}\s*\d{4,}.*$", t, re.IGNORECASE) and len(t) < 25:
        return True
    return False


def splitta_attivita(testo):
    """Divide il testo in attività elementari."""
    if not testo:
        return []
    # Split per + e newline
    parti = re.split(r'\+|\n', testo)
    risultato = []
    for parte in parti:
        parte = parte.strip()
        if parte and len(parte) > 2:
            risultato.append(parte)
    return risultato if risultato else [testo]


def classifica(testo):
    """Classifica un testo in una o più categorie. Ritorna lista di categorie."""
    if not testo or len(testo.strip()) < 3:
        return ["NON CLASSIFICATO"]

    # Se è solo un codice seriale, non classificabile
    if is_solo_codice(testo):
        return ["NON CLASSIFICATO"]

    categorie_trovate = []
    for nome, pattern in CATEGORIE_COMPILED:
        if pattern.search(testo):
            categorie_trovate.append(nome)

    return categorie_trovate if categorie_trovate else ["NON CLASSIFICATO"]


def categorizza_riga(descrizione, dettaglio):
    """Classifica una singola riga (descrizione + dettaglio) nelle macro-categorie.

    Args:
        descrizione: testo del campo Descrizione
        dettaglio: testo del campo Dettaglio

    Returns:
        Lista ordinata di categorie (es. ["01. PNEUMATICI"]),
        oppure ["NON CLASSIFICATO"] se nessuna categoria trovata.
    """
    descrizione = pulisci_testo(descrizione)
    dettaglio = pulisci_testo(dettaglio)

    testo_completo = f"{descrizione} {dettaglio}".strip() if dettaglio else descrizione

    # 1) Classifica sul testo intero combinato
    categorie_riga = set()
    for c in classifica(testo_completo):
        categorie_riga.add(c)

    # 2) Splitta e classifica ogni frammento
    for att in splitta_attivita(testo_completo):
        for c in classifica(att):
            categorie_riga.add(c)

    # Rimuovi NON CLASSIFICATO se almeno una categoria reale trovata
    if len(categorie_riga) > 1:
        categorie_riga.discard("NON CLASSIFICATO")

    return sorted(categorie_riga)


def rileva_colonne(ws):
    """Rileva automaticamente le colonne Descrizione e Dettaglio dall'header."""
    header = [str(cell.value).strip().upper() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_desc = None
    col_det = None

    # Mapping nomi noti
    nomi_desc = ["DESCRIZIONE", "MVDESART"]
    nomi_det = ["DETTAGLIO", "MVDESSUP"]

    for idx, nome in enumerate(header):
        if nome in nomi_desc:
            col_desc = idx
        elif nome in nomi_det:
            col_det = idx

    # Fallback: se non trovate, usa colonne 0 e 1
    if col_desc is None:
        col_desc = 0
    if col_det is None:
        col_det = 1

    # Colonne extra da includere nel dettaglio output
    extra_cols = {}
    for idx, nome in enumerate(header):
        if idx not in (col_desc, col_det) and nome:
            extra_cols[idx] = nome

    print(f"  Colonna Descrizione: {header[col_desc]} (col {col_desc + 1})")
    print(f"  Colonna Dettaglio:   {header[col_det]} (col {col_det + 1})")
    print(f"  Colonne extra:       {len(extra_cols)}")

    return col_desc, col_det, extra_cols, header


def main():
    print("Lettura manutenzione.xlsx...")
    wb_in = openpyxl.load_workbook("manutenzione.xlsx")
    ws_in = wb_in.active

    # Rileva colonne automaticamente
    col_desc, col_det, extra_cols, header = rileva_colonne(ws_in)

    # Raccolta dati
    catalogo = defaultdict(lambda: defaultdict(int))  # cat -> attività -> count
    dettaglio_righe = []  # per foglio 2

    for i, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True)):
        descrizione_raw = row[col_desc] if len(row) > col_desc and row[col_desc] else ""
        dettaglio_raw = row[col_det] if len(row) > col_det and row[col_det] else ""

        # Classificazione tramite funzione riutilizzabile
        categorie_lista = categorizza_riga(descrizione_raw, dettaglio_raw)
        categorie_str = " | ".join(categorie_lista)

        # Popola catalogo (serve il testo pulito per le sotto-attività)
        descrizione = pulisci_testo(descrizione_raw)
        dettaglio = pulisci_testo(dettaglio_raw)
        testo_completo = f"{descrizione} {dettaglio}".strip() if dettaglio else descrizione

        attivita_list = splitta_attivita(testo_completo)
        for att in attivita_list:
            cats = classifica(att)
            for c in cats:
                att_norm = re.sub(r'\s+', ' ', att.upper().strip())
                if len(att_norm) > 3:
                    catalogo[c][att_norm] += 1

        cats_solo_intero = set(classifica(testo_completo)) - {"NON CLASSIFICATO"}
        for c in cats_solo_intero:
            att_norm = re.sub(r'\s+', ' ', testo_completo.upper().strip())
            if len(att_norm) > 3:
                catalogo[c][att_norm] += 1

        # Dati riga
        riga_data = {
            "riga": i + 2,
            "descrizione": str(row[col_desc]).strip() if len(row) > col_desc and row[col_desc] else "",
            "dettaglio": str(row[col_det]).strip() if len(row) > col_det and row[col_det] else "",
            "categorie": categorie_str,
        }
        # Aggiungi colonne extra
        for idx, nome in extra_cols.items():
            val = row[idx] if len(row) > idx and row[idx] else ""
            riga_data[nome] = str(val).strip() if val else ""

        dettaglio_righe.append(riga_data)

    wb_in.close()

    # ── Creazione output Excel ────────────────────────────────────────────────
    print(f"Righe processate: {len(dettaglio_righe)}")

    wb_out = openpyxl.Workbook()

    # ── Stili ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    cat_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # ══════════════════════════════════════════════════════════════════════════
    # FOGLIO 1 – CATALOGO
    # ══════════════════════════════════════════════════════════════════════════
    ws_cat = wb_out.active
    ws_cat.title = "Catalogo"
    ws_cat.append(["Macro-Categoria", "Sotto-Attività", "Frequenza", "Esempi dal Dataset"])

    # Header style
    for col in range(1, 5):
        cell = ws_cat.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_num = 2
    # Ordina categorie
    categorie_ordinate = sorted(catalogo.keys())

    # Statistiche
    total_classificate = 0
    total_non = 0

    for cat in categorie_ordinate:
        attivita = catalogo[cat]
        # Raggruppa attività simili e ordina per frequenza
        attivita_ordinate = sorted(attivita.items(), key=lambda x: -x[1])

        # Riga intestazione categoria
        freq_totale = sum(v for _, v in attivita_ordinate)
        if cat == "NON CLASSIFICATO":
            total_non = freq_totale
        else:
            total_classificate += freq_totale

        ws_cat.cell(row=row_num, column=1, value=cat).font = cat_font
        ws_cat.cell(row=row_num, column=1).fill = cat_fill
        ws_cat.cell(row=row_num, column=3, value=freq_totale).font = cat_font
        ws_cat.cell(row=row_num, column=3).fill = cat_fill
        for col in range(1, 5):
            ws_cat.cell(row=row_num, column=col).border = thin_border
            ws_cat.cell(row=row_num, column=col).fill = cat_fill
        row_num += 1

        # Sotto-attività (max 50 per categoria per leggibilità)
        for att_nome, att_freq in attivita_ordinate[:50]:
            ws_cat.cell(row=row_num, column=2, value=att_nome).alignment = wrap_alignment
            ws_cat.cell(row=row_num, column=3, value=att_freq)
            for col in range(1, 5):
                ws_cat.cell(row=row_num, column=col).border = thin_border
            row_num += 1

    # Colonne auto-width
    ws_cat.column_dimensions["A"].width = 40
    ws_cat.column_dimensions["B"].width = 80
    ws_cat.column_dimensions["C"].width = 12
    ws_cat.column_dimensions["D"].width = 40

    # Freeze panes
    ws_cat.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════════════
    # FOGLIO 2 – DETTAGLIO (ogni riga originale → categorie)
    # ══════════════════════════════════════════════════════════════════════════
    ws_det = wb_out.create_sheet("Dettaglio")

    # Costruisci header dinamico: colonne extra + Riga, Descrizione, Dettaglio, Categorie
    extra_nomi = sorted(extra_cols.values())
    det_header = ["Riga Orig."] + extra_nomi + ["Descrizione", "Dettaglio", "Categorie Assegnate"]
    n_det_cols = len(det_header)
    col_cat_idx = n_det_cols  # 1-based index della colonna categorie

    ws_det.append(det_header)

    for col in range(1, n_det_cols + 1):
        cell = ws_det.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for r in dettaglio_righe:
        row_data = [r["riga"]]
        for nome in extra_nomi:
            row_data.append(r.get(nome, ""))
        row_data += [r["descrizione"], r["dettaglio"], r["categorie"]]
        ws_det.append(row_data)

    # Evidenzia NON CLASSIFICATO in rosso
    non_class_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row_idx in range(2, len(dettaglio_righe) + 2):
        cell = ws_det.cell(row=row_idx, column=col_cat_idx)
        if "NON CLASSIFICATO" in str(cell.value):
            cell.fill = non_class_fill
        for col in range(1, n_det_cols + 1):
            ws_det.cell(row=row_idx, column=col).border = thin_border
            ws_det.cell(row=row_idx, column=col).alignment = wrap_alignment

    ws_det.column_dimensions["A"].width = 10
    # Larghezza colonne extra
    for ci in range(2, 2 + len(extra_nomi)):
        ws_det.column_dimensions[get_column_letter(ci)].width = 18
    ws_det.column_dimensions[get_column_letter(2 + len(extra_nomi))].width = 45
    ws_det.column_dimensions[get_column_letter(3 + len(extra_nomi))].width = 80
    ws_det.column_dimensions[get_column_letter(4 + len(extra_nomi))].width = 50
    ws_det.freeze_panes = "A2"
    # Auto-filter
    ws_det.auto_filter.ref = f"A1:{get_column_letter(n_det_cols)}{len(dettaglio_righe) + 1}"

    # ══════════════════════════════════════════════════════════════════════════
    # FOGLIO 3 – RIEPILOGO
    # ══════════════════════════════════════════════════════════════════════════
    ws_riep = wb_out.create_sheet("Riepilogo")
    ws_riep.append(["Categoria", "N. Occorrenze", "% sul Totale"])

    for col in range(1, 4):
        cell = ws_riep.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    totale_righe = len(dettaglio_righe)
    # Conta righe per categoria (una riga può avere più categorie)
    conteggio_cat = defaultdict(int)
    for r in dettaglio_righe:
        for cat in r["categorie"].split(" | "):
            cat = cat.strip()
            if cat:
                conteggio_cat[cat] += 1

    row_num = 2
    for cat in sorted(conteggio_cat.keys()):
        n = conteggio_cat[cat]
        pct = n / totale_righe * 100
        ws_riep.cell(row=row_num, column=1, value=cat).border = thin_border
        ws_riep.cell(row=row_num, column=2, value=n).border = thin_border
        cell_pct = ws_riep.cell(row=row_num, column=3, value=round(pct, 1))
        cell_pct.border = thin_border
        cell_pct.number_format = '0.0"%"'
        row_num += 1

    # Riga totale
    ws_riep.cell(row=row_num, column=1, value="TOTALE RIGHE").font = Font(bold=True)
    ws_riep.cell(row=row_num, column=2, value=totale_righe).font = Font(bold=True)
    for col in range(1, 4):
        ws_riep.cell(row=row_num, column=col).border = thin_border

    ws_riep.column_dimensions["A"].width = 45
    ws_riep.column_dimensions["B"].width = 18
    ws_riep.column_dimensions["C"].width = 14
    ws_riep.freeze_panes = "A2"

    # Salva
    output_file = "catalogo_interventi.xlsx"
    wb_out.save(output_file)
    print(f"\nFile salvato: {output_file}")
    print(f"\n── Riepilogo ──")
    print(f"Righe totali: {totale_righe}")
    for cat in sorted(conteggio_cat.keys()):
        n = conteggio_cat[cat]
        print(f"  {cat}: {n} ({n/totale_righe*100:.1f}%)")


if __name__ == "__main__":
    main()
